#!/usr/bin/env python3
"""
Программа для поиска и скачивания фото товаров из Excel-файла прихода.

Читает .xls/.xlsx файл, парсит столбец "Товар" (формат: "Бренд Модель Цвет ..."),
ищет фотографии товаров через Яндекс.Картинки (Selenium) и сохраняет в папку images/.

Использование:
    python search_images.py приход6.xls
    python search_images.py приход6.xls --output ./my_images --count 3
"""

import argparse
import os
import re
import sys
import time
import logging
import json
from pathlib import Path
from dataclasses import dataclass
from urllib.parse import quote_plus

import requests
from PIL import Image
from io import BytesIO

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ─── Модели данных ────────────────────────────────────────────────────────────

@dataclass
class Product:
    """Товар, извлечённый из Excel."""
    row_number: int
    raw_name: str
    brand: str
    model: str
    color: str

    @property
    def search_query(self) -> str:
        """Поисковый запрос для нахождения фото."""
        return f"{self.brand} {self.model} {self.color} glasses"

    @property
    def filename(self) -> str:
        """Имя файла для сохранения (без расширения)."""
        safe = re.sub(r'[^\w\-]', '_', f"{self.brand}_{self.model}_{self.color}")
        return safe

    def __str__(self):
        return f"{self.brand} {self.model} (цвет: {self.color})"


# ─── Парсинг Excel ───────────────────────────────────────────────────────────

def read_xls(filepath: str) -> list[list[str]]:
    """Читает .xls файл (старый формат) и возвращает все строки."""
    import xlrd
    workbook = xlrd.open_workbook(filepath)
    sheet = workbook.sheet_by_index(0)
    rows = []
    for row_idx in range(sheet.nrows):
        row = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            row.append(str(cell.value).strip())
        rows.append(row)
    return rows


def read_xlsx(filepath: str) -> list[list[str]]:
    """Читает .xlsx файл (новый формат) и возвращает все строки."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows():
        cells = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
        rows.append(cells)
    wb.close()
    return rows


def read_excel(filepath: str) -> list[list[str]]:
    """Автоматически определяет формат и читает Excel файл."""
    ext = Path(filepath).suffix.lower()
    if ext == ".xls":
        return read_xls(filepath)
    elif ext in (".xlsx", ".xlsm"):
        return read_xlsx(filepath)
    else:
        # Пробуем оба формата
        try:
            return read_xls(filepath)
        except Exception:
            return read_xlsx(filepath)


def find_header_row(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    """
    Ищет строку-заголовок в данных. Возвращает индекс строки и словарь
    {нормализованное_имя: индекс_столбца}.
    
    Ищет строку, в которой есть НЕСКОЛЬКО ключевых слов-заголовков,
    чтобы не путать с текстом типа «Приход товаров №6».
    """
    # Ключевые слова — ищем точное совпадение ячейки (после strip/lower)
    exact_keywords = {"товар", "код", "наименование", "название", "№",
                      "количество", "цена", "сумма", "цена розн.", "артикул"}

    best_row = None
    best_score = 0

    for row_idx, row in enumerate(rows):
        non_empty = [cell.lower().strip() for cell in row if cell.strip()]
        score = 0
        for cell_text in non_empty:
            for kw in exact_keywords:
                if cell_text == kw:
                    score += 2  # Точное совпадение — высокий вес
                    break
                elif cell_text == kw + "." or kw == cell_text.rstrip("."):
                    score += 2
                    break

        # Нужно минимум 2 совпадения, чтобы считать строку заголовком
        if score >= 4 and score > best_score:
            best_score = score
            best_row = row_idx

    if best_row is not None:
        row = rows[best_row]
        header_map = {}
        for ci, cell in enumerate(row):
            if cell.strip():
                header_map[cell.lower().strip()] = ci
        log.info(f"Заголовок найден в строке {best_row + 1}: "
                 f"{[c for c in row if c.strip()]}")
        return best_row, header_map

    raise ValueError(
        "Не удалось найти строку-заголовок. Убедитесь, что в файле есть столбцы "
        "'Товар', 'Код', 'Количество', 'Цена' и т.д."
    )


def find_product_column(header_map: dict[str, int]) -> int:
    """Находит индекс столбца с названием товара."""
    # Приоритетные имена столбца
    priority = ["товар", "наименование", "название", "модель", "продукт", "артикул"]
    for name in priority:
        for key, idx in header_map.items():
            if name in key:
                log.info(f"Столбец с товарами: '{key}' (индекс {idx})")
                return idx
    raise ValueError(
        f"Не найден столбец с товарами. Доступные столбцы: {list(header_map.keys())}"
    )


def parse_product_name(raw: str) -> tuple[str, str, str] | None:
    """
    Парсит строку товара формата "Бренд МодельНомер Цвет [суффикс]".
    Примеры:
        "Chopard 358V 700 м/о" → ("Chopard", "358V", "700")
        "Ray-Ban RB3025 001/58" → ("Ray-Ban", "RB3025", "001/58")
        "Gucci GG0036S 002"     → ("Gucci", "GG0036S", "002")
    
    Возвращает (brand, model, color) или None если не распарсилось.
    """
    raw = raw.strip()
    if not raw:
        return None

    # Убираем суффиксы типа "м/о", "ж/о", "унисекс" и т.п.
    cleaned = re.sub(r'\s+[мжМЖ]/[оОcС]\s*$', '', raw)
    cleaned = re.sub(r'\s+унисекс\s*$', '', cleaned, flags=re.IGNORECASE)
    cleaned = cleaned.strip()

    # Разбиваем по пробелам
    parts = cleaned.split()
    if len(parts) < 2:
        return None

    # Бренд — первое слово (или несколько слов если бренд составной, напр. "Ray-Ban")
    # Модель — обычно содержит цифры
    # Цвет — последний числовой/буквенно-числовой блок

    brand = parts[0]
    
    # Если первая часть — только буквы и вторая тоже только буквы, склеиваем бренд
    # (для случаев типа "Tom Ford TF5401 052")
    idx = 1
    while idx < len(parts) and parts[idx].isalpha() and not any(c.isdigit() for c in parts[idx]):
        # Но проверяем — если следующий элемент выглядит как модель, останавливаемся
        if idx + 1 < len(parts):
            brand += " " + parts[idx]
            idx += 1
        else:
            break

    if idx >= len(parts):
        # Всё оказалось брендом — нет модели
        return None

    model = parts[idx]
    idx += 1

    # Цвет — то что осталось (может быть пустым)
    color = " ".join(parts[idx:]) if idx < len(parts) else ""

    return brand, model, color


def extract_products(filepath: str) -> list[Product]:
    """Извлекает список товаров из Excel файла."""
    rows = read_excel(filepath)
    if not rows:
        raise ValueError("Файл пуст!")

    header_idx, header_map = find_header_row(rows)
    product_col = find_product_column(header_map)

    products = []
    seen = set()

    for row_idx in range(header_idx + 1, len(rows)):
        row = rows[row_idx]
        if product_col >= len(row):
            continue

        raw_name = row[product_col].strip()
        if not raw_name or raw_name == "0.0" or raw_name == "None":
            continue

        parsed = parse_product_name(raw_name)
        if parsed is None:
            log.warning(f"  Строка {row_idx + 1}: не удалось распарсить '{raw_name}', пропуск")
            continue

        brand, model, color = parsed

        # Дедупликация — один и тот же товар может быть несколько раз
        key = f"{brand}|{model}|{color}".lower()
        if key in seen:
            log.debug(f"  Строка {row_idx + 1}: дубликат '{raw_name}', пропуск")
            continue
        seen.add(key)

        products.append(Product(
            row_number=row_idx + 1,
            raw_name=raw_name,
            brand=brand,
            model=model,
            color=color,
        ))

    return products


# ─── Selenium + Яндекс.Картинки ──────────────────────────────────────────────

_driver = None


def get_driver():
    """Создаёт и возвращает Selenium WebDriver (Chrome headless, singleton)."""
    global _driver
    if _driver is not None:
        return _driver

    log.info("Запуск Chrome (headless)...")
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    # Подавляем лишние логи Chrome
    opts.add_argument("--log-level=3")
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])

    _driver = webdriver.Chrome(options=opts)

    log.info("Chrome запущен.")
    return _driver


def close_driver():
    """Закрывает браузер."""
    global _driver
    if _driver is not None:
        try:
            _driver.quit()
        except Exception:
            pass
        _driver = None


def search_images_yandex(query: str, max_results: int = 8) -> list[str]:
    """
    Ищет изображения через Яндекс.Картинки и возвращает список URL
    полноразмерных изображений.
    """
    driver = get_driver()
    urls = []

    try:
        # Формируем URL поиска по Яндекс.Картинкам
        # isize=large — большие изображения, icolor=white — белый фон (по возможности)
        encoded_query = quote_plus(query)
        search_url = (
            f"https://yandex.ru/images/search"
            f"?text={encoded_query}"
            f"&isize=large"
        )

        log.debug(f"  URL: {search_url}")
        driver.get(search_url)

        # Ждём загрузки результатов
        time.sleep(2)

        # Способ 1: Извлекаем URL из data-атрибутов элементов
        # Яндекс хранит данные об изображениях в JSON внутри data-bem атрибутов
        selectors = [
            "a.serp-item__link",
            ".serp-item",
            ".serp-item__preview",
            "a.Link.ContentImage-Cover",
            ".ContentImage-Cover",
            ".justifier__item",
        ]

        items = []
        for selector in selectors:
            try:
                items = driver.find_elements(By.CSS_SELECTOR, selector)
                if items:
                    log.debug(f"  Найдено {len(items)} элементов по '{selector}'")
                    break
            except Exception:
                continue

        # Пробуем извлечь URL из data-bem (JSON с данными об изображении)
        for item in items[:max_results * 2]:
            try:
                data_bem = item.get_attribute("data-bem")
                if data_bem:
                    data = json.loads(data_bem)
                    # Внутри может быть serp-item → img_href или dups[0].url
                    serp = data.get("serp-item", {})
                    img_url = serp.get("img_href") or serp.get("preview", [{}])[0].get("url", "")
                    if img_url and img_url.startswith("http"):
                        urls.append(img_url)
                        if len(urls) >= max_results:
                            break
                    # dups — дубликаты, часто содержат оригинальную ссылку
                    for dup in serp.get("dups", []):
                        dup_url = dup.get("img_href") or dup.get("url", "")
                        if dup_url and dup_url.startswith("http"):
                            urls.append(dup_url)
                            if len(urls) >= max_results:
                                break
            except Exception:
                continue

        # Способ 2: Если не получилось через data-bem, пробуем кликать на картинки
        if not urls:
            log.debug("  Пробуем извлечь URL через просмотр картинок...")
            try:
                # Кликаем первую картинку для открытия просмотра
                img_links = driver.find_elements(By.CSS_SELECTOR, "a[href*='/images/search']")
                if not img_links:
                    img_links = driver.find_elements(By.CSS_SELECTOR, ".serp-item img, .ContentImage img")
                
                for link in img_links[:max_results]:
                    try:
                        link.click()
                        time.sleep(1)
                        # В режиме просмотра ищем оригинальную ссылку
                        orig_links = driver.find_elements(
                            By.CSS_SELECTOR,
                            "a.MMViewerButtons-OpenImage, a[class*='Origin'], a[href*='img_url']"
                        )
                        for orig in orig_links:
                            href = orig.get_attribute("href")
                            if href and "img_url=" in href:
                                # Извлекаем img_url из параметра
                                import urllib.parse
                                parsed = urllib.parse.urlparse(href)
                                params = urllib.parse.parse_qs(parsed.query)
                                if "img_url" in params:
                                    urls.append(params["img_url"][0])
                            elif href and href.startswith("http") and "yandex" not in href:
                                urls.append(href)

                        # Закрываем просмотр
                        try:
                            close_btn = driver.find_element(
                                By.CSS_SELECTOR, ".MMViewerModal-Close, .Modal-Close"
                            )
                            close_btn.click()
                            time.sleep(0.3)
                        except Exception:
                            driver.back()
                            time.sleep(0.5)

                    except Exception:
                        continue

            except Exception as e:
                log.debug(f"  Ошибка при клике: {e}")

        # Способ 3: Извлекаем все src больших картинок
        if not urls:
            log.debug("  Пробуем извлечь src из img-тегов...")
            try:
                imgs = driver.find_elements(By.CSS_SELECTOR, "img.serp-item__thumb, img.ContentImage-Image")
                for img_el in imgs[:max_results]:
                    src = img_el.get_attribute("src")
                    if src and src.startswith("http"):
                        # Яндекс проксирует через свой CDN — пробуем использовать как есть
                        urls.append(src)
            except Exception:
                pass

        # Способ 4: Через page source парсим JSON из скрипта
        if not urls:
            log.debug("  Пробуем извлечь URL из page source...")
            try:
                page_source = driver.page_source
                # Ищем все URL изображений в исходнике
                img_pattern = re.findall(
                    r'"img_href"\s*:\s*"(https?://[^"]+)"', page_source
                )
                urls.extend(img_pattern[:max_results])
            except Exception:
                pass

    except Exception as e:
        log.error(f"Ошибка Яндекс поиска для '{query}': {e}")

    # Убираем дубликаты, сохраняя порядок
    seen = set()
    unique_urls = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            unique_urls.append(u)

    log.debug(f"  Найдено {len(unique_urls)} URL")
    return unique_urls[:max_results]


def download_image(url: str, timeout: int = 15) -> bytes | None:
    """Скачивает изображение по URL. Возвращает байты или None."""
    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        }
        resp = requests.get(url, headers=headers, timeout=timeout, stream=True)
        resp.raise_for_status()

        # Проверяем Content-Type
        content_type = resp.headers.get("Content-Type", "")
        if "image" not in content_type and "octet" not in content_type:
            return None

        data = resp.content
        # Проверяем, что это валидное изображение
        img = Image.open(BytesIO(data))
        img.verify()

        # Проверяем минимальный размер (чтобы не скачивать миниатюры)
        img = Image.open(BytesIO(data))
        w, h = img.size
        if w < 200 or h < 200:
            log.debug(f"  Слишком маленькое ({w}x{h}), пропуск: {url[:80]}")
            return None

        return data

    except Exception as e:
        log.debug(f"  Ошибка скачивания {url[:80]}: {e}")
        return None


def get_image_extension(data: bytes) -> str:
    """Определяет расширение файла по содержимому."""
    try:
        img = Image.open(BytesIO(data))
        fmt = img.format
        if fmt:
            return fmt.lower().replace("jpeg", "jpg")
    except Exception:
        pass
    return "jpg"


def search_and_download(product: Product, output_dir: str, max_images: int = 1) -> list[str]:
    """
    Ищет и скачивает изображения для товара.
    Возвращает список путей к скачанным файлам.
    """
    saved = []

    # Пробуем несколько вариантов поискового запроса
    queries = [
        f"{product.brand} {product.model} {product.color} eyewear white background",
        f"{product.brand} {product.model} {product.color}",
        f"{product.brand} {product.model} glasses",
    ]

    for query in queries:
        if len(saved) >= max_images:
            break

        log.info(f"  Поиск (Яндекс): '{query}'")
        urls = search_images_yandex(query, max_results=8)

        if not urls:
            log.warning(f"  Нет результатов для '{query}'")
            continue

        for url in urls:
            if len(saved) >= max_images:
                break

            data = download_image(url)
            if data is None:
                continue

            ext = get_image_extension(data)
            idx_suffix = f"_{len(saved) + 1}" if max_images > 1 else ""
            filename = f"{product.filename}{idx_suffix}.{ext}"
            filepath = os.path.join(output_dir, filename)

            with open(filepath, "wb") as f:
                f.write(data)

            saved.append(filepath)
            img = Image.open(BytesIO(data))
            log.info(f"  ✓ Сохранено: {filename} ({img.size[0]}x{img.size[1]})")

        if saved:
            break  # Нашли по текущему запросу — не пробуем следующие

    return saved


# ─── Главная функция ──────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Поиск и скачивание фото товаров из Excel-файла прихода",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Примеры:
  python search_images.py приход6.xls
  python search_images.py приход6.xls --output ./фото --count 3
  python search_images.py приход6.xls --delay 2

Формат файла:
  Программа ищет столбец "Товар" (или "Наименование") и парсит строки
  формата "Бренд Модель Цвет", например:
    Chopard 358V 700 м/о  →  бренд=Chopard, модель=358V, цвет=700
        """,
    )
    parser.add_argument("input", help="Путь к Excel-файлу (.xls или .xlsx)")
    parser.add_argument(
        "--output", "-o",
        default="images",
        help="Папка для сохранения изображений (по умолч. 'images')",
    )
    parser.add_argument(
        "--count", "-c",
        type=int,
        default=1,
        help="Количество фото на каждый товар (по умолч. 1)",
    )
    parser.add_argument(
        "--delay", "-d",
        type=float,
        default=1.0,
        help="Задержка между запросами в секундах (по умолч. 1.0)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Только показать найденные товары, не скачивать",
    )

    args = parser.parse_args()

    # Проверяем входной файл
    if not os.path.isfile(args.input):
        log.error(f"Файл не найден: {args.input}")
        sys.exit(1)

    # Извлекаем товары
    log.info(f"Чтение файла: {args.input}")
    products = extract_products(args.input)

    if not products:
        log.error("Товары не найдены в файле!")
        sys.exit(1)

    log.info(f"Найдено товаров: {len(products)}")
    print()

    # Показываем список
    for i, p in enumerate(products, 1):
        print(f"  {i:3d}. {p}")
    print()

    if args.dry_run:
        log.info("Режим dry-run — скачивание пропущено.")
        return

    # Создаём папку для результатов
    os.makedirs(args.output, exist_ok=True)

    # Скачиваем
    success = 0
    failed = 0

    for i, product in enumerate(products, 1):
        log.info(f"[{i}/{len(products)}] {product}")
        saved = search_and_download(product, args.output, max_images=args.count)

        if saved:
            success += 1
        else:
            failed += 1
            log.warning(f"  ✗ Не удалось найти фото для {product}")

        # Задержка между запросами
        if i < len(products):
            time.sleep(args.delay)

    # Закрываем браузер
    close_driver()

    # Итоги
    print()
    log.info("=" * 50)
    log.info(f"Готово! Успешно: {success}, Не найдено: {failed}")
    log.info(f"Фото сохранены в: {os.path.abspath(args.output)}")


if __name__ == "__main__":
    try:
        main()
    finally:
        close_driver()
